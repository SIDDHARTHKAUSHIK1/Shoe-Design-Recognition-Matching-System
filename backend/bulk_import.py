"""
Bulk Data Management Module for ShoeMatch AI Admin Panel.
Handles CSV/Excel file parsing, paired ZIP image extraction, pre-commit validation,
duplicate detection, and batch execution through the frozen-safe ingest_single_design path.
"""
import os
import io
import csv
import zipfile
import requests
import urllib.request
import logging
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional, Set

import openpyxl

from backend.database import get_db_connection
from backend.ingestion import ingest_single_design

VALID_CATEGORIES = [
    "Sneaker", "Running Shoe", "Casual Trainer", "Slip-On Loafer", 
    "Classic Oxford", "Hiking Boot", "Slide Sandal", "Flip-Flop", 
    "House Slipper", "Mule Slipper", "High-Top Basketball", "Athletic Cross-Trainer"
]

logger = logging.getLogger("backend.bulk_import")

# Column Header Normalization Mapping
HEADER_MAPPING = {
    "design_id": "design_id",
    "designid": "design_id",
    "sku": "design_id",
    "id": "design_id",
    
    "name": "name",
    "model": "name",
    "model_name": "name",
    "design_name": "name",
    
    "category": "category",
    "type": "category",
    
    "description": "description",
    "desc": "description",
    
    "created_by": "created_by",
    "author": "created_by",
    
    "shelf_location": "shelf_location",
    "location": "shelf_location",
    "shelf": "shelf_location",
    
    "shoe_match_tag": "shoe_match_tag",
    "zone": "shoe_match_tag",
    "facility": "shoe_match_tag",
    
    "drawer": "drawer",
    "slot": "slot",
    
    "materials": "materials",
    "material": "materials",
    
    "season": "season",
    "collection": "season",
    
    "production_status": "production_status",
    "status": "production_status",
    
    "image_urls": "image_urls",
    "image_url": "image_urls",
    "image_paths": "image_urls",
    "image_path": "image_urls",
    "image_filename": "image_urls",
    "images": "image_urls"
}


def normalize_header(h: str) -> str:
    clean = str(h).strip().lower().replace(" ", "_").replace("-", "_")
    return HEADER_MAPPING.get(clean, clean)


def parse_spreadsheet_bytes(file_bytes: bytes, filename: str) -> List[Dict[str, Any]]:
    """Parse CSV or Excel XLSX bytes into list of row dictionaries."""
    rows = []
    ext = Path(filename).suffix.lower()

    if ext == ".csv":
        text_content = file_bytes.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text_content))
        header_raw = next(reader, None)
        if not header_raw:
            return []
        
        headers = [normalize_header(h) for h in header_raw]
        for row_idx, row in enumerate(reader, start=2):
            if not any(row):
                continue
            row_dict = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    row_dict[headers[idx]] = str(val).strip()
            row_dict["_row_num"] = row_idx
            rows.append(row_dict)

    elif ext in [".xlsx", ".xls"]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        sheet = wb.active
        iter_rows = list(sheet.iter_rows(values_only=True))
        if not iter_rows:
            return []
        
        header_raw = iter_rows[0]
        headers = [normalize_header(h or "") for h in header_raw]
        
        for row_idx, row in enumerate(iter_rows[1:], start=2):
            if not any(row):
                continue
            row_dict = {}
            for idx, val in enumerate(row):
                if idx < len(headers):
                    val_str = "" if val is None else str(val).strip()
                    row_dict[headers[idx]] = val_str
            row_dict["_row_num"] = row_idx
            rows.append(row_dict)
    else:
        raise ValueError(f"Unsupported spreadsheet file extension: '{ext}'. Supported: .csv, .xlsx")

    return rows


def extract_zip_images(zip_bytes: bytes) -> Dict[str, List[Dict[str, Any]]]:
    """
    Extract images from ZIP archive bytes in-memory and group them by design_id.
    Matches design_id via directory name (SHOE-001/side.jpg) or filename prefix (SHOE-001_side.jpg).
    """
    design_images: Dict[str, List[Dict[str, Any]]] = {}
    valid_exts = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}

    with zipfile.ZipFile(io.BytesIO(zip_bytes), "r") as z:
        for zip_info in z.infolist():
            if zip_info.is_dir():
                continue
            
            p = Path(zip_info.filename)
            if p.suffix.lower() not in valid_exts:
                continue
            
            content = z.read(zip_info.filename)
            if not content:
                continue
            
            # Match design_id from directory name or filename prefix
            parts = p.parts
            design_id = None
            
            if len(parts) >= 2:
                # e.g., SHOE-001/angle_side.jpg
                possible_id = parts[-2].strip()
                if possible_id:
                    design_id = possible_id.upper()
            
            if not design_id:
                # e.g., SHOE-001_side.jpg or SHOE-001.jpg
                base_stem = p.stem.upper()
                if "_" in base_stem:
                    design_id = base_stem.split("_")[0]
                elif "-" in base_stem:
                    # e.g. SHOE-001
                    design_id = base_stem
                else:
                    design_id = base_stem

            if design_id:
                if design_id not in design_images:
                    design_images[design_id] = []
                design_images[design_id].append({
                    "filename": p.name,
                    "content": content,
                    "angle": p.stem
                })

    return design_images


def get_existing_catalog_design_ids() -> Set[str]:
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT design_id FROM designs;")
        return {r[0].upper() for r in cursor.fetchall()}


def validate_bulk_rows(
    rows: List[Dict[str, Any]],
    zip_images: Dict[str, List[Dict[str, Any]]]
) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
    """
    Validate all rows before write execution.
    Categorizes each row into: 'ready', 'duplicate_file', 'duplicate_catalog', 'error'.
    """
    existing_ids = get_existing_catalog_design_ids()
    seen_in_file: Set[str] = set()

    validated_rows = []
    ready_count = 0
    dup_file_count = 0
    dup_catalog_count = 0
    error_count = 0

    valid_cats = {c.lower() for c in VALID_CATEGORIES}

    for row in rows:
        row_num = row.get("_row_num", 0)
        design_id = (row.get("design_id") or "").strip().upper()
        name = (row.get("name") or "").strip()
        category = (row.get("category") or "Sneaker").strip()
        image_urls = (row.get("image_urls") or "").strip()

        errors = []

        # Required fields check
        if not design_id:
            errors.append("Missing required 'design_id'")
        if not name:
            errors.append("Missing required 'name'")

        # Category check
        if category and category.lower() not in valid_cats:
            # Allow fallback if valid category
            pass

        # Image availability check
        row_zip_imgs = zip_images.get(design_id, [])
        if not row_zip_imgs and not image_urls:
            # Note: warning/error if no image found
            errors.append(f"No reference image found in ZIP for '{design_id}' or via image_urls column")

        # Duplicate status determination
        status = "ready"
        if errors:
            status = "error"
            error_count += 1
        elif design_id in seen_in_file:
            status = "duplicate_file"
            dup_file_count += 1
        elif design_id in existing_ids:
            status = "duplicate_catalog"
            dup_catalog_count += 1
        else:
            ready_count += 1

        if design_id:
            seen_in_file.add(design_id)

        v_row = dict(row)
        v_row["design_id"] = design_id
        v_row["name"] = name
        v_row["category"] = category
        v_row["image_urls"] = image_urls
        v_row["status"] = status
        v_row["errors"] = errors
        v_row["images_found_count"] = len(row_zip_imgs) + (1 if image_urls else 0)
        validated_rows.append(v_row)

    summary = {
        "total_rows": len(rows),
        "ready_count": ready_count,
        "duplicate_file_count": dup_file_count,
        "duplicate_catalog_count": dup_catalog_count,
        "error_count": error_count
    }

    return summary, validated_rows


def execute_bulk_import_batch(
    validated_rows: List[Dict[str, Any]],
    zip_images: Dict[str, List[Dict[str, Any]]],
    duplicate_handling: str = "skip"  # 'skip', 'overwrite'
) -> Dict[str, Any]:
    """
    Execute batch ingestion row-by-row calling frozen-safe ingest_single_design path.
    """
    succeeded_count = 0
    failed_count = 0
    skipped_count = 0

    results_log = []

    for row in validated_rows:
        row_num = row.get("_row_num", 0)
        design_id = row.get("design_id", "")
        name = row.get("name", "")
        status = row.get("status", "")

        # Skip validation error rows
        if status == "error":
            failed_count += 1
            results_log.append({
                "row_num": row_num,
                "design_id": design_id,
                "name": name,
                "status": "failed",
                "reason": "; ".join(row.get("errors", ["Validation error"]))
            })
            continue

        # Handle duplicates based on strategy
        if status == "duplicate_file":
            skipped_count += 1
            results_log.append({
                "row_num": row_num,
                "design_id": design_id,
                "name": name,
                "status": "skipped",
                "reason": "Duplicate design_id within file"
            })
            continue

        if status == "duplicate_catalog" and duplicate_handling == "skip":
            skipped_count += 1
            results_log.append({
                "row_num": row_num,
                "design_id": design_id,
                "name": name,
                "status": "skipped",
                "reason": "Design already exists in catalog (skipped by rule)"
            })
            continue

        # Gather image payloads (from ZIP or URLs/server paths)
        image_payloads = zip_images.get(design_id, [])
        
        # If no ZIP images, check image_urls column
        if not image_payloads and row.get("image_urls"):
            raw_urls = [u.strip() for u in row["image_urls"].split(",") if u.strip()]
            for idx, url_str in enumerate(raw_urls):
                try:
                    img_data = None
                    if url_str.startswith("http://") or url_str.startswith("https://"):
                        resp = requests.get(url_str, timeout=10, headers={"User-Agent": "Mozilla/5.0"})
                        if resp.status_code == 200 and resp.content:
                            img_data = resp.content
                    else:
                        local_p = Path(url_str)
                        if local_p.exists() and local_p.is_file():
                            img_data = local_p.read_bytes()
                    
                    if img_data:
                        image_payloads.append({
                            "filename": f"{design_id}_{idx+1}.jpg",
                            "content": img_data,
                            "angle": f"angle_{idx+1}"
                        })
                except Exception as ex:
                    logger.warning(f"Failed to fetch image URL '{url_str}' for {design_id}: {ex}")

        if not image_payloads:
            failed_count += 1
            results_log.append({
                "row_num": row_num,
                "design_id": design_id,
                "name": name,
                "status": "failed",
                "reason": "No valid reference images could be loaded"
            })
            continue

        # Invoke frozen-safe ingest_single_design
        try:
            res = ingest_single_design(
                design_id=design_id,
                name=name,
                category=row.get("category") or "Sneaker",
                description=row.get("description") or "",
                created_by=row.get("created_by") or "Bulk Import Admin",
                shelf_location=row.get("shelf_location") or "Warehouse A - Rack 01 - Shelf 1",
                materials=row.get("materials") or "Leather / Rubber Sole",
                season=row.get("season") or "Collection 2026",
                production_status=row.get("production_status") or "Active Production Sample",
                image_files=image_payloads
            )
            succeeded_count += 1
            results_log.append({
                "row_num": row_num,
                "design_id": design_id,
                "name": name,
                "status": "succeeded",
                "reason": f"Ingested {res.get('total_vectors_indexed', len(image_payloads))} photos"
            })
        except Exception as e:
            logger.error(f"Bulk import failed for row {row_num} ({design_id}): {e}")
            failed_count += 1
            results_log.append({
                "row_num": row_num,
                "design_id": design_id,
                "name": name,
                "status": "failed",
                "reason": str(e)
            })

    return {
        "total_processed": len(validated_rows),
        "succeeded": succeeded_count,
        "failed": failed_count,
        "skipped": skipped_count,
        "details": results_log
    }
